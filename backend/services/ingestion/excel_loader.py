from openpyxl import load_workbook


def load_excel(file_path: str, source_id: str, filename: str) -> list[dict]:
    """Extract rows from Excel sheets, grouping every 20 rows into a chunk."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []

        for row in ws.iter_rows(values_only=True):
            cell_values = [str(cell) if cell is not None else "" for cell in row]
            row_text = " | ".join(cell_values).strip()
            if row_text and row_text != " | ".join([""] * len(cell_values)):
                rows_text.append(row_text)

        # Group every 20 rows into one chunk
        group_size = 20
        for i in range(0, len(rows_text), group_size):
            group = rows_text[i : i + group_size]
            chunk_text = "\n".join(group)
            if chunk_text.strip():
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source_id": source_id,
                        "filename": filename,
                        "source_type": "excel",
                        "sheet_name": sheet_name,
                        "row_range": f"{i + 1}-{i + len(group)}",
                    },
                })

    wb.close()
    return chunks
